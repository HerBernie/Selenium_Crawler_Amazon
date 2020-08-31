#!/usr/bin/env python
# coding=utf-8
'''
@project: Selenium_Crawler_Amazon
@author: Boyang Xia (Alvin)
@file: main.py
@time: 2020/8/18 16:54
@desc:
'''

import SkuListGenerator
import reviewsCrawler
# import time
# import pandas as pd
import json
import queue
import threading
import openpyxl

# load configurations
merchantInfo = json.load(open('merchantInfo.json', mode='r'))

def sku_crawler():
    skuCrawlerConfig = json.load(open('skuCrawlerConfig.json', mode='r'))

    def a_proxy():
        global proxyPointer
        if proxyPointer == len(skuCrawlerConfig['proxyServers']):
            proxyPointer = 0
        proxy = skuCrawlerConfig['proxyServers'][proxyPointer]
        proxyPointer += 1
        return proxy

    threadIdList = [id for id in range(skuCrawlerConfig['threadNum'])]
    dateList = [skuCrawlerConfig['year']+skuCrawlerConfig['month']+str(date).rjust(2, '0') for date in range(int(skuCrawlerConfig['startDate']), int(skuCrawlerConfig['endDate'])+1)]
    # dateList = ['190406', '190408']
    dateQueue = queue.Queue(1 + int(skuCrawlerConfig['endDate']) - int(skuCrawlerConfig['startDate']))
    queueLock = threading.Lock()

    queueLock.acquire()
    for formatDate in dateList:
        dateQueue.put(formatDate)
    queueLock.release()

    proxyPointer = 0

    # main loop
    while not dateQueue.empty():
        threads = []

        for threadId in threadIdList:
            thread = SkuListGenerator.SkuListGeneratorThread(threadId, dateQueue, queueLock, skuCrawlerConfig['searchRange'], merchantInfo, a_proxy())
            thread.start()
            threads.append(thread)
        for thread in threads:
            thread.join()

            # open spreadsheet workbook
            try:
                wb = openpyxl.load_workbook(f"{merchantInfo['name']}_sku_list.xlsx")
            except FileNotFoundError:
                wb = openpyxl.workbook.Workbook()
            try:
                sheet = wb[f"{skuCrawlerConfig['year']}{skuCrawlerConfig['month']}"]
            except KeyError:
                sheet = wb.create_sheet(f"{skuCrawlerConfig['year']}{skuCrawlerConfig['month']}")
                sheet.append(['SKU', 'LINK', 'ASIN', 'TITLE', 'PRICE', 'STOCK'])

            for record in thread.skuRecordList:
                sheet.append(record)

            del thread
            wb.save(f"{merchantInfo['name']}_sku_list.xlsx")

    print('exit, writen to xlsx.')

def reviews_crawler():
    aCrawler = reviewsCrawler.ReviewsCrawler(merchantInfo)
    aCrawler.start()
    if input() == '0':
        aCrawler.exitFlag = 1


exitFlag = 0

while exitFlag == 0:
    print('[1]: skuCrawler')
    print('[2]: reviewsCrawler')
    str = input('enter[1/2](0 to exit): ')
    if str == '1':
        sku_crawler()
    elif str == '2':
        reviews_crawler()
    elif str == '0':
        exitFlag = 1
    else:
        print('error')





